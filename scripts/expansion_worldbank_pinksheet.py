#!/usr/bin/env python3
"""
World Bank Commodity Price (Pink Sheet) Collector
=================================================
Monthly World Bank CMO (Commodity Markets Outlook) "Pink Sheet" data into
data/macro/commodities_monthly.csv.

Schema:
    date,iron_ore,coal_newcastle,coal_australian,crude_brent,natgas_eu,natgas_us,
    lng_japan,gasoline,diesel,fertilizer_dap,maize,wheat_srw,rice_thai,soybeans,
    copper,aluminum,nickel,
    base_metals,precious_metals,energy_index,agri_index,total_index

Notes (verified):
- The static task URL serves a stale vintage; the CMO landing page is scraped
  for the current monthly xlsx and the freshest candidate wins.
- No Newcastle series in current vintages: coal_newcastle column stays empty.
- Monthly dates normalized to YYYY-MM-01. Missing = empty string, never 0.
"""

import csv
import io
import os
import re
import sys
import time

import requests

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_PATH = os.path.join(REPO_ROOT, 'data', 'macro', 'commodities_monthly.csv')

FIELDNAMES = ['date', 'iron_ore', 'coal_newcastle', 'coal_australian', 'crude_brent',
              'natgas_eu', 'natgas_us', 'lng_japan', 'gasoline', 'diesel',
              'fertilizer_dap', 'maize', 'wheat_srw', 'rice_thai', 'soybeans',
              'copper', 'aluminum', 'nickel',
              'base_metals', 'precious_metals', 'energy_index', 'agri_index', 'total_index']

BROWSER_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36',
}

CMO_LANDING = "https://www.worldbank.org/en/research/commodity-markets"
FALLBACK_URLS = [
    "https://thedocs.worldbank.org/en/doc/5d903e848db1d1b83e0ec8f744e55570-0350012021/related/CMO-Historical-Data-Monthly.xlsx",
]

# Pink Sheet column header fragments -> slug
COMMODITY_MAP = {
    'iron ore': 'iron_ore',
    'newcastle': 'coal_newcastle',
    'coal, australian': 'coal_australian',
    'crude oil, brent': 'crude_brent',
    'natural gas, europe': 'natgas_eu',
    'natural gas, u.s.': 'natgas_us',
    'liquefied natural gas, japan': 'lng_japan',
    'gasoline': 'gasoline',
    'diesel': 'diesel',
    'dap': 'fertilizer_dap',
    'maize': 'maize',
    'corn': 'maize',
    'wheat, us srw': 'wheat_srw',
    'rice, thai 5%': 'rice_thai',
    'soybeans': 'soybeans',
    'copper': 'copper',
    'aluminum': 'aluminum',
    'nickel': 'nickel',
}

INDEX_MAP = {
    'base metals': 'base_metals',
    'precious metals': 'precious_metals',
    'energy': 'energy_index',
    'agriculture': 'agri_index',
    'total index': 'total_index',
}

RETRIES = 3
TIMEOUT = 30


def fetch_with_retry(url):
    for attempt in range(1, RETRIES + 1):
        try:
            r = requests.get(url, headers=BROWSER_HEADERS, timeout=TIMEOUT)
            if r.status_code == 200:
                return r
            print(f"[warn] HTTP {r.status_code} for {url[:90]} (attempt {attempt}/{RETRIES})", file=sys.stderr)
        except requests.RequestException as e:
            print(f"[warn] {e} (attempt {attempt}/{RETRIES})", file=sys.stderr)
        time.sleep(2 ** attempt)
    return None


def resolve_current_xlsx():
    """Scrape the CMO landing page for the current monthly data xlsx."""
    resp = fetch_with_retry(CMO_LANDING)
    candidates = []
    if resp is not None:
        for m in re.finditer(r'href="([^"]+\.xlsx)"', resp.text, re.I):
            candidates.append(m.group(1))
    candidates = [u if u.startswith('http') else 'https://www.worldbank.org' + u for u in candidates]
    # Monthly historical file preferred
    monthly = [u for u in candidates if 'monthly' in u.lower()] or candidates
    urls = monthly + FALLBACK_URLS
    for url in urls:
        resp = fetch_with_retry(url)
        if resp is not None and len(resp.content) > 10000:
            print(f"[ok] using xlsx: {url[:110]}")
            return resp.content
    return None


def normalize_month(v):
    """Excel serial or datetime or '2026M07' -> 'YYYY-MM-01'."""
    import openpyxl.utils.datetime as _dt
    try:
        if hasattr(v, 'year') and hasattr(v, 'month'):
            return f"{v.year}-{v.month:02d}-01"
        if isinstance(v, (int, float)) and v > 20000:
            base = _dt.from_excel(v)
            return f"{base.year}-{base.month:02d}-01"
        s = str(v).strip()
        m = re.match(r'^(\d{4})M(\d{1,2})$', s)
        if m:
            return f"{m.group(1)}-{int(m.group(2)):02d}-01"
        from datetime import datetime
        dt = datetime.strptime(s, '%Y-%m-%d')
        return f"{dt.year}-{dt.month:02d}-01"
    except Exception:
        return None


def parse_workbook(content):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    # 'Monthly Prices' sheet: header rows ~6-7, data from row 8+
    sheet = None
    for name in ('Monthly Prices', 'Monthly Sheet', 'Monthly'):
        if name in wb.sheetnames:
            sheet = wb[name]
            break
    if sheet is None:
        return {}

    rows = list(sheet.iter_rows(values_only=True))
    # Locate the header row containing 'Date'/month markers and commodity names
    header_row_idx = None
    headers = []
    for i, row in enumerate(rows[:12]):
        cells = [str(c).strip().lower() if c is not None else '' for c in row]
        joined = ' '.join(cells)
        if 'crude' in joined or 'iron ore' in joined:
            header_row_idx = i
            headers = cells
            break
    if header_row_idx is None:
        return {}

    # Some vintages stack two header rows; merge non-empty fragments upward
    col_slug = {}
    for j, h in enumerate(headers):
        if not h:
            continue
        for frag, slug in COMMODITY_MAP.items():
            if frag in h:
                col_slug[j] = slug
                break

    data = {}
    for row in rows[header_row_idx + 1:]:
        if not row or row[0] is None:
            continue
        iso = normalize_month(row[0])
        if not iso:
            continue
        rec = {}
        for j, slug in col_slug.items():
            if slug is None or j >= len(row):
                continue
            v = row[j]
            try:
                if v is not None and str(v).strip() != '':
                    rec[slug] = round(float(v), 6)
            except (ValueError, TypeError):
                continue
        if rec:
            data[iso] = rec
    return data


def _norm_header(h):
    s = re.sub(r'\*+', '', str(h or '')).strip().lower()
    return re.sub(r'\s+', ' ', s)

def parse_indices_workbook(content):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    name = next((n for n in wb.sheetnames if 'indice' in n.lower() or 'index' in n.lower()), None)
    if name is None:
        return {}
    sheet = wb[name]
    rows = list(sheet.iter_rows(values_only=True))

    # Stacked multi-row header: collect slug mapping by scanning the top block
    SLUGS = {'total index': 'total_index', 'energy': 'energy_index',
             'base metals': 'base_metals', 'precious metals': 'precious_metals',
             'agriculture': 'agri_index'}
    col_slug = {}
    for i, row in enumerate(rows[:12]):
        if i >= len(rows):
            break
        for j, h in enumerate(row):
            nh = _norm_header(h)
            if nh in SLUGS and j not in col_slug:
                col_slug[j] = SLUGS[nh]

    def is_data_row(row):
        return bool(row) and row[0] is not None and re.match(r'^(\d{4})M(\d{1,2})$', str(row[0]).strip())

    data = {}
    for row in rows:
        if not is_data_row(row):
            continue
        iso = normalize_month(row[0])
        if not iso:
            continue
        rec = {}
        for j, slug in col_slug.items():
            if j >= len(row):
                continue
            v = row[j]
            try:
                if v is not None and str(v).strip() != '':
                    rec[slug] = round(float(v), 6)
            except (ValueError, TypeError):
                continue
        if rec:
            data[iso] = rec
    return data


def main():
    content = resolve_current_xlsx()
    if content is None:
        print("[FAIL] could not retrieve any CMO xlsx", file=sys.stderr)
        return 1

    existing = {}
    if os.path.exists(OUT_PATH):
        with open(OUT_PATH, encoding='utf-8', newline='') as f:
            for row in csv.DictReader(f):
                existing[row['date']] = row
    before = len(existing)

    prices = parse_workbook(content)
    indices = parse_indices_workbook(content)
    print(f"parsed months: prices={len(prices)}, indices={len(indices)}")

    all_dates = sorted(set(existing) | set(prices) | set(indices))
    for d in all_dates:
        row = existing.get(d, {'date': d})
        for slug, val in prices.get(d, {}).items():
            if not row.get(slug):
                row[slug] = val
        for slug, val in indices.get(d, {}).items():
            if not row.get(slug):
                row[slug] = val
        existing[d] = row

    rows = [{k: r.get(k, '') or '' for k in FIELDNAMES} for r in (existing[d] for d in sorted(existing))]
    # Drop series that the Pink Sheet no longer publishes (all-empty columns,
    # e.g. coal_newcastle/gasoline after WB's 2026 series restructure) so the
    # CSV only carries fields that actually contain data.
    live_fields = [FIELDNAMES[0]] + [
        k for k in FIELDNAMES[1:] if any(r[k].strip() for r in rows)
    ]
    if len(live_fields) < len(FIELDNAMES):
        dropped = [k for k in FIELDNAMES if k not in live_fields]
        print(f"[info] dropping empty columns: {', '.join(dropped)}")
        rows = [{k: r[k] for k in live_fields} for r in rows]
    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    tmp = OUT_PATH + '.tmp'
    with open(tmp, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=live_fields)
        writer.writeheader()
        writer.writerows(rows)
    import shutil as _sh
    _sh.move(tmp, OUT_PATH)
    print(f"[ok] commodities_monthly.csv: {before} -> {len(rows)} months")
    return 0


if __name__ == '__main__':
    sys.exit(main())
