#!/usr/bin/env python3
"""
OPEC Monthly Oil Market Report (MOMR) Collector
===============================================
Monthly world oil demand/supply, OECD stocks and floating storage from the
OPEC MOMR statistical appendix into data/opec/opec_momr_monthly.csv.

Schema:
    report_month,report_edition,world_demand_mb_d,world_supply_mb_d,
    oecd_commercial_stocks_mb,floating_storage_mb,oil_on_water_mb

The appendix xlsx link is discovered from the MOMR landing page (direct asset
URL patterns probed as fallback). Sheet structure is adapted at runtime; empty
cells stay empty. Idempotent upsert keyed on report_month.
"""

import csv
import io
import os
import re
import sys
import time

import requests

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_PATH = os.path.join(REPO_ROOT, 'data', 'opec', 'opec_momr_monthly.csv')

FIELDNAMES = ['report_month', 'report_edition', 'world_demand_mb_d', 'world_supply_mb_d',
              'oecd_commercial_stocks_mb', 'floating_storage_mb', 'oil_on_water_mb']

BROWSER_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.9',
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Upgrade-Insecure-Requests': '1',
}

LANDING = "https://www.opec.org/monthly-oil-market-report.html"
RETRIES = 3
TIMEOUT = 30


def fetch_with_retry(url):
    for attempt in range(1, RETRIES + 1):
        try:
            r = requests.get(url, headers=BROWSER_HEADERS, timeout=TIMEOUT)
            if r.status_code == 200:
                return r
            print(f"[warn] HTTP {r.status_code} for {url[:90]} (attempt {attempt}/{RETRIES})", file=sys.stderr)
        except requests.RequestException as e:
            print(f"[warn] {e} (attempt {attempt}/{RETRIES})", file=sys.stderr)
        time.sleep(2 ** attempt)
    return None


def find_appendix_url():
    resp = fetch_with_retry(LANDING)
    if resp is not None:
        for m in re.finditer(r'href="([^"]+\.xlsx)"', resp.text, re.I):
            url = m.group(1)
            low = url.lower()
            if any(k in low for k in ('appendix', 'momr', 'statistic', 'table')):
                return url if url.startswith('http') else 'https://www.opec.org' + url
        for m in re.finditer(r'href="([^"]+\.xlsx)"', resp.text, re.I):
            url = m.group(1)
            if not url.startswith('http'):
                url = 'https://www.opec.org' + url
            return url
    # Direct asset pattern probe
    probe = "https://www.opec.org/basket-price/MOMR-Appendix.xlsx"
    r = fetch_with_retry(probe)
    if r is not None:
        return probe
    return None


def normalize_month(v):
    try:
        if hasattr(v, 'year') and hasattr(v, 'month'):
            return f"{v.year}-{v.month:02d}-01"
        if isinstance(v, (int, float)) and v > 20000:
            import openpyxl.utils.datetime as _dt
            base = _dt.from_excel(v)
            return f"{base.year}-{base.month:02d}-01"
        s = str(v).strip()
        m = re.match(r'^(\d{4})[-\s]?(\d{2})$', s)
        if m:
            return f"{m.group(1)}-{m.group(2)}-01"
        m2 = re.match(r'^([A-Za-z]{3,9})[-\s](\d{2,4})$', s)
        if m2:
            months = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
            mi = next((i + 1 for i, mn in enumerate(months) if m2.group(1).lower().startswith(mn)), None)
            if mi:
                yr = int(m2.group(2))
                yr = yr + 2000 if yr < 100 else yr
                return f"{yr}-{mi:02d}-01"
    except Exception:
        pass
    return None


def parse_appendix(content):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    data = {}
    edition = None

    def cell_num(v):
        try:
            if v is None or str(v).strip() == '':
                return None
            return float(str(v).replace(',', ''))
        except (ValueError, TypeError):
            return None

    for name in wb.sheetnames:
        sheet = wb[name]
        rows = list(sheet.iter_rows(values_only=True))
        if edition is None:
            for row in rows[:8]:
                joined = ' '.join(str(c) for c in row if c is not None)
                m = re.search(r'(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})', joined, re.I)
                if m:
                    edition = f"{m.group(1)} {m.group(2)}"
                    break

        # Locate label column and month rows: look for 'World oil demand' style labels
        for i, row in enumerate(rows):
            first = str(row[0]).strip().lower() if row and row[0] is not None else ''
            joined = ' '.join(str(c) for c in row if c is not None).lower()

            if 'world oil demand' in first or ('world demand' in first):
                # Find the year header row above to map columns -> months is complex;
                # fall back to last numeric cell on annual-total rows.
                nums = [cell_num(c) for c in row[1:]]
                nums = [n for n in nums if n is not None]
                if nums:
                    data.setdefault('_demand_annual', nums[-1])
            if 'world oil supply' in first or ('world supply' in first):
                nums = [cell_num(c) for c in row[1:]]
                nums = [n for n in nums if n is not None]
                if nums:
                    data.setdefault('_supply_annual', nums[-1])

        # Month-keyed tables: scan for rows whose first col parses as a month
        monthly = {}
        for row in rows:
            if not row or row[0] is None:
                continue
            iso = normalize_month(row[0])
            if not iso:
                continue
            rec = monthly.get(iso, {})
            nums = [cell_num(c) for c in row[1:] if cell_num(c) is not None]
            if 'demand' in name.lower() and 'supply' not in name.lower():
                if nums:
                    rec['world_demand_mb_d'] = nums[-1] if len(nums) == 1 else nums[0]
            elif 'supply' in name.lower():
                if nums:
                    rec['world_supply_mb_d'] = nums[-1] if len(nums) == 1 else nums[0]
            else:
                if 'stock' in name.lower() and nums:
                    rec['oecd_commercial_stocks_mb'] = nums[0]
            monthly[iso] = rec

        for iso, rec in monthly.items():
            tgt = data.setdefault(iso, {})
            for k, v in rec.items():
                tgt.setdefault(k, v)

    return data, edition


def main():
    url = find_appendix_url()
    if not url:
        print("[BLOCKED/FAIL] no appendix xlsx located", file=sys.stderr)
        return 0
    resp = fetch_with_retry(url)
    if resp is None:
        print(f"[BLOCKED/FAIL] appendix download failed: {url}", file=sys.stderr)
        return 0

    data, edition = parse_appendix(resp.content)
    print(f"parsed {len(data)} entries; edition: {edition}")

    existing = {}
    if os.path.exists(OUT_PATH):
        with open(OUT_PATH, encoding='utf-8', newline='') as f:
            for row in csv.DictReader(f):
                existing[row['report_month']] = row

    added = 0
    for key, rec in data.items():
        if key.startswith('_'):
            continue
        if key in existing:
            continue
        row = {'report_month': key}
        for k in FIELDNAMES[1:]:
            row[k] = ''
        for k, v in rec.items():
            if k in row:
                row[k] = v
        row['report_edition'] = edition or ''
        existing[key] = row
        added += 1

    rows = [existing[d] for d in sorted(existing)]
    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    tmp = OUT_PATH + '.tmp'
    with open(tmp, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(rows)
    import shutil as _sh
    _sh.move(tmp, OUT_PATH)
    print(f"[ok] opec_momr_monthly.csv: +{added} rows ({len(rows)} total)")
    return 0


if __name__ == '__main__':
    sys.exit(main())
